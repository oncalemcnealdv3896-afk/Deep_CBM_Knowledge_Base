from __future__ import annotations

import csv
import json
import re
import string
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "00_Master" / "Deep_CBM_Latest.xlsx"
INDEX_DIR = ROOT / "00_Master" / "index"
STATS_DIR = ROOT / "06_Statistics"
EXPECTED_FIRST_SHEET = "完整文献矩阵"

EXPORT_FIELDS = [
    "文献ID",
    "DOI",
    "DOI_normalized",
    "英文题名",
    "英文题名_normalized",
    "中文题名",
    "中文题名_normalized",
    "第一作者",
    "年份",
    "期刊/来源",
    "官方网页",
    "核验状态",
]

SOURCE_FIELDS = [
    "文献ID",
    "DOI",
    "英文题名",
    "中文题名",
    "第一作者",
    "年份",
    "期刊/来源",
    "官方网页",
    "核验状态",
]

ALIASES = {
    "文献ID": ["文献ID", "文献编号", "ID", "Literature ID", "Record ID"],
    "DOI": ["DOI", "doi"],
    "英文题名": ["英文题名", "英文标题", "Title", "English Title", "Article Title"],
    "中文题名": ["中文题名", "中文标题", "Chinese Title"],
    "第一作者": ["第一作者", "First Author", "Author", "Authors"],
    "年份": ["年份", "Year", "Publication Year", "发表年份"],
    "期刊/来源": ["期刊/来源", "期刊", "来源", "Journal", "Source", "Publication"],
    "官方网页": ["官方网页", "URL", "url", "链接", "Official URL", "网页"],
    "核验状态": ["核验状态", "验证状态", "状态", "Verification Status", "QC Status"],
}

PUNCT_TRANSLATION = str.maketrans("", "", string.punctuation + "，。、《》？?！!：:；;（）()【】[]“”‘’'\"·—–-_")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_doi(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"^https?://(dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi\s*:\s*", "", text, flags=re.IGNORECASE)
    return text.strip()


def normalize_title(value: Any) -> str:
    text = normalize_spaces(clean_text(value)).lower()
    text = text.translate(PUNCT_TRANSLATION)
    return normalize_spaces(text)


def normalize_header(value: Any) -> str:
    return normalize_spaces(clean_text(value)).lower()


def header_score(values: list[Any]) -> int:
    normalized = {normalize_header(value) for value in values if clean_text(value)}
    score = 0
    for aliases in ALIASES.values():
        if any(normalize_header(alias) in normalized for alias in aliases):
            score += 1
    return score


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    candidate_rows = [1] + list(range(2, min(ws.max_row, 20) + 1))
    best_row = 1
    best_score = -1
    for row_idx in candidate_rows:
        values = [cell.value for cell in ws[row_idx]]
        score = header_score(values)
        if score > best_score:
            best_row = row_idx
            best_score = score
        if row_idx == 1 and score >= 3:
            break

    header_map: dict[str, int] = {}
    row_values = [cell.value for cell in ws[best_row]]
    normalized_headers = {normalize_header(value): idx + 1 for idx, value in enumerate(row_values) if clean_text(value)}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            key = normalize_header(alias)
            if key in normalized_headers:
                header_map[field] = normalized_headers[key]
                break
    return best_row, header_map


def dcbm_number(value: Any) -> int | None:
    match = re.search(r"DCBM[-_\s]?(\d+)", clean_text(value), flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def duplicate_count(values: list[str]) -> int:
    counts = Counter(value for value in values if value)
    return sum(count - 1 for count in counts.values() if count > 1)


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def main() -> int:
    if not SOURCE.exists():
        print(f"Source workbook not found: {SOURCE}", file=sys.stderr)
        return 1

    wb = load_workbook(SOURCE, read_only=True, data_only=False)
    try:
        first_sheet_name = wb.sheetnames[0]
        if first_sheet_name != EXPECTED_FIRST_SHEET:
            print(
                f"First worksheet must be {EXPECTED_FIRST_SHEET}, got {first_sheet_name}",
                file=sys.stderr,
            )
            return 2

        ws = wb[first_sheet_name]
        header_row, header_map = find_header_row(ws)
        missing_fields = [field for field in SOURCE_FIELDS if field not in header_map]

        records: list[dict[str, str]] = []
        dcbm_numbers: list[int] = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            source_values: dict[str, str] = {}
            for field in SOURCE_FIELDS:
                col = header_map.get(field)
                source_values[field] = clean_text(row[col - 1].value) if col else ""

            if not any(source_values.values()):
                continue

            record = {
                "文献ID": source_values["文献ID"],
                "DOI": source_values["DOI"],
                "DOI_normalized": normalize_doi(source_values["DOI"]),
                "英文题名": source_values["英文题名"],
                "英文题名_normalized": normalize_title(source_values["英文题名"]),
                "中文题名": source_values["中文题名"],
                "中文题名_normalized": normalize_title(source_values["中文题名"]),
                "第一作者": source_values["第一作者"],
                "年份": source_values["年份"],
                "期刊/来源": source_values["期刊/来源"],
                "官方网页": source_values["官方网页"],
                "核验状态": source_values["核验状态"],
            }
            number = dcbm_number(record["文献ID"])
            if number is not None:
                dcbm_numbers.append(number)
            records.append(record)

        ids_continuous = False
        last_dcbm = "未识别"
        if dcbm_numbers:
            unique_numbers = sorted(set(dcbm_numbers))
            ids_continuous = unique_numbers == list(range(unique_numbers[0], unique_numbers[-1] + 1)) and len(unique_numbers) == len(dcbm_numbers)
            last_dcbm = f"DCBM-{unique_numbers[-1]:03d}"

        INDEX_DIR.mkdir(parents=True, exist_ok=True)
        STATS_DIR.mkdir(parents=True, exist_ok=True)
        csv_path = INDEX_DIR / "Deep_CBM_master_index.csv"
        jsonl_path = INDEX_DIR / "Deep_CBM_master_index.jsonl"
        today = date.today().strftime("%Y%m%d")
        report_path = STATS_DIR / f"Master_Index_Report_{today}.md"

        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=EXPORT_FIELDS)
            writer.writeheader()
            writer.writerows(records)

        with jsonl_path.open("w", encoding="utf-8", newline="\n") as handle:
            for record in records:
                handle.write(json.dumps(record, ensure_ascii=False) + "\n")

        doi_missing = sum(1 for record in records if not record["DOI_normalized"])
        doi_duplicates = duplicate_count([record["DOI_normalized"] for record in records])
        english_duplicates = duplicate_count([record["英文题名_normalized"] for record in records])
        chinese_duplicates = duplicate_count([record["中文题名_normalized"] for record in records])

        report = f"""# Master Index Report

- 运行日期：{date.today().isoformat()}
- 源文件路径：`{rel(SOURCE)}`
- 第一张 sheet 名称：{first_sheet_name}
- 表头行：{header_row}
- 总记录数：{len(records)}
- 最后一个 DCBM 编号：{last_dcbm}
- 文献ID是否连续：{'是' if ids_continuous else '否'}
- DOI 缺失数量：{doi_missing}
- DOI 重复数量：{doi_duplicates}
- 英文题名重复数量：{english_duplicates}
- 中文题名重复数量：{chinese_duplicates}
- 缺失字段列表：{', '.join(missing_fields) if missing_fields else '无'}

## 输出文件路径

- `{rel(csv_path)}`
- `{rel(jsonl_path)}`
- `{rel(report_path)}`
"""
        report_path.write_text(report, encoding="utf-8", newline="\n")

        print(f"Wrote {rel(csv_path)}")
        print(f"Wrote {rel(jsonl_path)}")
        print(f"Wrote {rel(report_path)}")
        return 0
    finally:
        wb.close()


if __name__ == "__main__":
    raise SystemExit(main())