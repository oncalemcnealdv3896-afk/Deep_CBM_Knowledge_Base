"""Master workbook updater — append new records while preserving all sheets, formats, formulas."""

from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
MASTER = ROOT / "00_Master" / "Deep_CBM_Latest.xlsx"
SNAPSHOTS = ROOT / "00_Master" / "snapshots"

SKIP_VALUES = {"待定", "TBD", "待全文核验", ""}

SHEET_MATRIX = "完整文献矩阵"
SHEET_STATS = "总览统计"
SHEET_WEB = "网页索引"
SHEET_FIELDS = "字段说明"
SHEET_SUMMARY = "精读摘要"
SHEET_CHANGELOG = "更新日志"
SHEET_DUPCHECK = "重复检查"
SHEET_QC = "QC报告"

ALL_SHEETS = [
    SHEET_MATRIX, SHEET_STATS, SHEET_WEB, SHEET_FIELDS,
    SHEET_SUMMARY, SHEET_CHANGELOG, SHEET_DUPCHECK, SHEET_QC,
]

# Master matrix columns (1-indexed)
COL_MAP = {
    "序号": 1, "文献ID": 2, "第一作者": 3, "全部作者": 4, "年份": 5,
    "英文题名": 6, "中文题名": 7, "期刊/来源": 8, "卷": 9, "期": 10,
    "页码/文章号": 11, "DOI": 12, "官方网页": 13, "文献类型": 14,
    "研究区域/区块": 15, "研究对象": 16, "埋深/对象深度": 17,
    "研究目的": 18, "主要方法": 19, "核心结论": 20, "创新点": 21,
    "局限性": 22, "与你课题的关联": 23, "建议引用章节": 24,
    "关键词/标签": 25, "精读优先级": 26, "核验状态": 27,
}

FIELD_TO_COL = {
    "provisional_id": "文献ID",
    "first_author": "第一作者",
    "authors": "全部作者",
    "year": "年份",
    "title_en": "英文题名",
    "title_zh": "中文题名",
    "journal": "期刊/来源",
    "volume": "卷",
    "issue": "期",
    "pages_or_article": "页码/文章号",
    "doi": "DOI",
    "official_url": "官方网页",
    "document_type": "文献类型",
    "study_area": "研究区域/区块",
    "research_object": "研究对象",
    "burial_depth": "埋深/对象深度",
    "research_purpose": "研究目的",
    "methods": "主要方法",
    "findings": "核心结论",
    "innovation": "创新点",
    "limitations": "局限性",
    "relevance": "与你课题的关联",
    "suggested_chapter": "建议引用章节",
    "tags": "关键词/标签",
    "reading_priority": "精读优先级",
    "verification_status": "核验状态",
}


def create_snapshot() -> Path:
    """Copy current master workbook to snapshots dir, return snapshot path."""
    today = date.today().strftime("%Y-%m-%d")
    snapshot = SNAPSHOTS / f"Deep_CBM_Knowledge_Base_{today}.xlsx"
    # If already exists today, add a suffix
    if snapshot.exists():
        snapshot = SNAPSHOTS / f"Deep_CBM_Knowledge_Base_{today}_daily_brief_update.xlsx"
    SNAPSHOTS.mkdir(parents=True, exist_ok=True)
    shutil.copy2(MASTER, snapshot)
    return snapshot


def apply_new_records(
    candidates: list[dict[str, Any]],
    proposed_ids: list[str],
    duplicate_log: list[dict[str, Any]],
    run_dir: Path,
    dry_run: bool = False,
) -> Path:
    """Append new records to master workbook, update all auxiliary sheets.

    Args:
        candidates: New candidate dicts (must be "new" classification).
        proposed_ids: Final DCBM IDs for each candidate.
        duplicate_log: Duplicate entries for the duplicate check sheet.
        run_dir: Run directory for temp output.
        dry_run: If True, save to run_dir/tmp_workbook.xlsx instead of MASTER.

    Returns path to the saved workbook.
    """
    output = run_dir / "tmp_workbook.xlsx" if dry_run else MASTER
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(MASTER)

    # ---- 1. Append to 完整文献矩阵 ------------------------------------------
    ws = wb[SHEET_MATRIX]
    header_row = 1
    data_start = header_row + 1
    last_data_row = ws.max_row

    # Determine starting sequence number
    last_seq = 0
    for row_idx in range(data_start, last_data_row + 1):
        val = ws.cell(row=row_idx, column=COL_MAP["序号"]).value
        try:
            last_seq = max(last_seq, int(val))
        except (TypeError, ValueError):
            pass

    for i, (cand, final_id) in enumerate(zip(candidates, proposed_ids)):
        new_row = last_data_row + 1 + i
        seq = last_seq + 1 + i

        row_data = {
            "序号": seq,
            "文献ID": final_id,
            "第一作者": cand.get("first_author", ""),
            "全部作者": cand.get("authors", ""),
            "年份": cand.get("year", ""),
            "英文题名": cand.get("title_en", ""),
            "中文题名": cand.get("title_zh", ""),
            "期刊/来源": cand.get("journal", ""),
            "卷": cand.get("volume", ""),
            "期": cand.get("issue", ""),
            "页码/文章号": cand.get("pages_or_article", ""),
            "DOI": cand.get("doi", ""),
            "官方网页": cand.get("official_url", ""),
            "文献类型": cand.get("document_type", ""),
            "研究区域/区块": cand.get("study_area", ""),
            "研究对象": cand.get("research_object", ""),
            "埋深/对象深度": cand.get("burial_depth", ""),
            "研究目的": cand.get("research_purpose", ""),
            "主要方法": cand.get("methods", ""),
            "核心结论": cand.get("findings", ""),
            "创新点": cand.get("innovation", ""),
            "局限性": cand.get("limitations", ""),
            "与你课题的关联": cand.get("relevance", ""),
            "建议引用章节": cand.get("suggested_chapter", ""),
            "关键词/标签": cand.get("tags", ""),
            "精读优先级": cand.get("reading_priority", ""),
            "核验状态": cand.get("verification_status", ""),
        }

        # Copy formatting from the previous data row
        ref_row = last_data_row if last_data_row >= data_start else data_start
        for col_name, col_idx in COL_MAP.items():
            cell = ws.cell(row=new_row, column=col_idx)
            cell.value = row_data.get(col_name, "")
            _copy_style(ws.cell(row=ref_row, column=col_idx), cell)

        last_data_row = new_row

    # ---- 2. Update 总览统计 ------------------------------------------------
    _update_stats_sheet(wb, last_seq + len(candidates))

    # ---- 3. Update 网页索引 ------------------------------------------------
    _update_web_index(wb, candidates, proposed_ids, last_seq)

    # ---- 4. Update 精读摘要 ------------------------------------------------
    _update_summary_sheet(wb, candidates, proposed_ids)

    # ---- 5. Update 更新日志 ------------------------------------------------
    _update_changelog_sheet(wb, candidates, proposed_ids, duplicate_log)

    # ---- 6. Update 重复检查 ------------------------------------------------
    if duplicate_log:
        _update_dupcheck_sheet(wb, duplicate_log)

    # ---- 7. Update QC报告 --------------------------------------------------
    _update_qc_sheet(wb, len(candidates), len(duplicate_log))

    # ---- Save --------------------------------------------------------------
    wb.save(str(output))
    wb.close()
    return output


# ---------------------------------------------------------------------------
# Auxiliary sheet updaters
# ---------------------------------------------------------------------------

def _update_stats_sheet(wb, total_records: int) -> None:
    """Full update of 总览统计 — metadata rows + year/journal distributions."""
    ws = wb[SHEET_STATS]
    today_str = date.today().isoformat()

    from collections import Counter
    ws_matrix = wb[SHEET_MATRIX]
    years: Counter[str] = Counter()
    journals: Counter[str] = Counter()

    header_row = 1
    col_map = {}
    for cell in ws_matrix[header_row]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    for row in ws_matrix.iter_rows(min_row=header_row + 1):
        if not any(c.value for c in row):
            continue
        def _get(field: str) -> str:
            c = col_map.get(field)
            if c:
                for cell in row:
                    if cell.column == c:
                        return str(cell.value).strip() if cell.value else ""
            return ""
        y, j = _get("年份"), _get("期刊/来源")
        if y:
            years[y] += 1
        if j:
            journals[j] += 1

    # --- Update metadata rows by label (handle both old and Codex formats) ---
    ids = sorted([int(str(ws_matrix.cell(row=r, column=2).value or "").replace("DCBM-", "").replace("DCBM", ""))
                  for r in range(2, ws_matrix.max_row + 1)
                  if str(ws_matrix.cell(row=r, column=2).value or "").startswith("DCBM")])
    last_id = f"DCBM-{ids[-1]:03d}" if ids else "N/A"

    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if not label:
            continue
        # Match various label formats
        if "累计文献数" in label or "当前累计文献数" in label:
            ws.cell(row=row_idx, column=2).value = total_records
        elif "最后编号" in label or "当前最后编号" in label:
            ws.cell(row=row_idx, column=2).value = last_id
        elif "版本" in label or "当前版本" in label:
            ws.cell(row=row_idx, column=2).value = f"v1_{total_records}_daily_brief_update_{today_str}"
        elif "网页索引记录数" in label:
            ws.cell(row=row_idx, column=2).value = total_records
        elif "精读摘要记录数" in label:
            ws.cell(row=row_idx, column=2).value = total_records
        elif "更新" in label and ("最近" in label or "最后" in label or "日期" in label):
            ws.cell(row=row_idx, column=2).value = f"{today_str} Daily Brief Update"

    # --- Update year distribution ---
    for row_idx in range(10, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value or "")
        if "年份" in label and "分布" in label:
            # Find all existing year rows after this header
            r = row_idx + 1
            for y_year in sorted(years.keys()):
                while r < ws.max_row:
                    v = str(ws.cell(row=r, column=1).value or "").strip()
                    if v and v != y_year and not v.isdigit():
                        break
                    if v == y_year or (not v and r > row_idx + 1):
                        ws.cell(row=r, column=1).value = y_year
                        ws.cell(row=r, column=2).value = years[y_year]
                        r += 1
                        break
                    r += 1
            break

    # --- Update journal distribution ---
    for row_idx in range(20, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value or "")
        if "期刊" in label and ("来源" in label or "分布" in label):
            r = row_idx + 1
            sorted_j = sorted(journals.items(), key=lambda x: -x[1])
            for j_name, cnt in sorted_j:
                found = False
                for rr in range(row_idx + 1, ws.max_row + 1):
                    if str(ws.cell(row=rr, column=1).value or "").strip() == j_name:
                        ws.cell(row=rr, column=2).value = cnt
                        found = True
                        break
                if not found:
                    while r < ws.max_row and ws.cell(row=r, column=1).value:
                        r += 1
                    if r < ws.max_row:
                        ws.cell(row=r, column=1).value = j_name
                        ws.cell(row=r, column=2).value = cnt
            break

    # --- Update total record count row ---
    for row_idx in range(ws.max_row - 10, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value or "")
        if "总记录数" in label:
            ws.cell(row=row_idx, column=2).value = total_records
            ws.cell(row=row_idx, column=3).value = today_str
            break


def _update_web_index(
    wb, candidates: list[dict[str, Any]], proposed_ids: list[str], last_seq: int
) -> None:
    ws = wb[SHEET_WEB]
    last_row = ws.max_row
    for i, (cand, final_id) in enumerate(zip(candidates, proposed_ids)):
        new_row = last_row + 1 + i
        ws.cell(row=new_row, column=1).value = last_seq + 1 + i
        ws.cell(row=new_row, column=2).value = final_id
        ws.cell(row=new_row, column=3).value = cand.get("year", "")
        ws.cell(row=new_row, column=4).value = cand.get("title_en", "")
        ws.cell(row=new_row, column=5).value = cand.get("journal", "")
        ws.cell(row=new_row, column=6).value = cand.get("doi", "")
        ws.cell(row=new_row, column=7).value = cand.get("official_url", "")
        ws.cell(row=new_row, column=8).value = cand.get("verification_status", "")
        # Copy style from previous row
        ref_row = last_row if last_row >= 2 else 2
        for col in range(1, 9):
            _copy_style(ws.cell(row=ref_row, column=col), ws.cell(row=new_row, column=col))


def _update_summary_sheet(
    wb, candidates: list[dict[str, Any]], proposed_ids: list[str]
) -> None:
    """Update 精读摘要 with entries matching existing Codex format.

    Cols: A=文献ID B=年份 C=题名 D=研究区/对象 E=标准化精读摘要
          F=研究目的 G=数据/方法 H=核心发现 I=创新点 J=局限性
          K=与你课题关系 L=建议引用位置 M=精读优先级 N=摘要升级状态
    """
    ws = wb[SHEET_SUMMARY]
    last_row = ws.max_row

    for i, (cand, final_id) in enumerate(zip(candidates, proposed_ids)):
        new_row = last_row + 1 + i

        priority_raw = cand.get("reading_priority", "").strip().upper()
        star_map = {"A": "★★★★★", "B": "★★★★", "C": "★★★"}
        priority_stars = star_map.get(priority_raw, priority_raw)

        study_parts = []
        for p in [cand.get("study_area", ""), cand.get("research_object", "")]:
            if p and p not in SKIP_VALUES:
                study_parts.append(p)
        burial = cand.get("burial_depth", "")
        if burial and burial not in SKIP_VALUES:
            study_parts.append(f"深度：{burial}")
        study_str = "; ".join(study_parts) if study_parts else ""

        detail = cand.get("_detail_prose", {})
        summary_parts = []
        for tag, field in [
            ("研究目的", "research_purpose"),
            ("数据/方法", "methods"),
            ("核心发现", "findings"),
            ("创新点", "innovation"),
            ("局限性", "limitations"),
            ("与你课题的关系", "relevance"),
        ]:
            val = detail.get(f"{field}_prose", "") or cand.get(field, "")
            if val and val not in SKIP_VALUES:
                summary_parts.append(f"【{tag}】{val}")
        summary_str = "\n".join(summary_parts) if summary_parts else ""

        def _clean_val(field: str) -> str:
            val = cand.get(field, "")
            if "待全文核验" in val:
                prose = detail.get(f"{field}_prose", "")
                if prose:
                    return prose
            return val

        ws.cell(row=new_row, column=1).value = final_id
        ws.cell(row=new_row, column=2).value = cand.get("year", "")
        ws.cell(row=new_row, column=3).value = cand.get("title_zh", "")
        ws.cell(row=new_row, column=4).value = study_str
        ws.cell(row=new_row, column=5).value = summary_str
        ws.cell(row=new_row, column=6).value = _clean_val("research_purpose")
        ws.cell(row=new_row, column=7).value = _clean_val("methods")
        ws.cell(row=new_row, column=8).value = _clean_val("findings")
        ws.cell(row=new_row, column=9).value = _clean_val("innovation")
        ws.cell(row=new_row, column=10).value = _clean_val("limitations")
        ws.cell(row=new_row, column=11).value = _clean_val("relevance")
        ws.cell(row=new_row, column=12).value = _clean_val("suggested_chapter")
        ws.cell(row=new_row, column=13).value = priority_stars
        ws.cell(row=new_row, column=14).value = "由 Daily Brief 生成；待全文精读增强"


def _update_changelog_sheet(
    wb,
    candidates: list[dict[str, Any]],
    proposed_ids: list[str],
    duplicate_log: list[dict[str, Any]],
) -> None:
    ws = wb[SHEET_CHANGELOG]
    last_row = ws.max_row
    new_row = last_row + 1
    today_str = date.today().isoformat()
    added = ", ".join(proposed_ids) if proposed_ids else "无"
    skipped = len(duplicate_log)
    ws.cell(row=new_row, column=1).value = f"{today_str} Daily Brief Update"
    ws.cell(row=new_row, column=2).value = (
        f"新增: {added}; 跳过重复: {skipped}; "
        f"主库记录数更新为 {last_row} → (待QC确认)"
    )


def _update_dupcheck_sheet(wb, duplicate_log: list[dict[str, Any]]) -> None:
    ws = wb[SHEET_DUPCHECK]
    last_row = ws.max_row
    for i, entry in enumerate(duplicate_log):
        new_row = last_row + 1 + i
        ws.cell(row=new_row, column=1).value = entry.get("provisional_id", "")
        ws.cell(row=new_row, column=2).value = entry.get("existing_doi", "")
        ws.cell(row=new_row, column=3).value = entry.get("duplicate_reason", "")
        ws.cell(row=new_row, column=4).value = date.today().isoformat()
        ws.cell(row=new_row, column=5).value = entry.get("existing_dcbm_id", "")
        ws.cell(row=new_row, column=6).value = "已跳过"


def _update_qc_sheet(wb, added: int, skipped: int) -> None:
    ws = wb[SHEET_QC]
    last_row = ws.max_row
    new_row = last_row + 1
    ws.cell(row=new_row, column=1).value = f"Daily Brief Update {date.today().isoformat()}"
    ws.cell(row=new_row, column=2).value = f"新增 {added} 篇, 跳过重复 {skipped} 篇, 待 validate 确认"


# ---------------------------------------------------------------------------
# Style helper
# ---------------------------------------------------------------------------

def _copy_style(src, dst) -> None:
    """Copy cell style from src to dst (best-effort)."""
    if src.has_style:
        dst.font = src.font.copy()
        dst.border = src.border.copy()
        dst.fill = src.fill.copy()
        dst.number_format = src.number_format
        dst.protection = src.protection.copy()
        dst.alignment = src.alignment.copy()
