"""Index exporter — regenerate machine-readable master index from workbook."""

from __future__ import annotations

import csv
import json
import re
import string
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
MASTER = ROOT / "00_Master" / "Deep_CBM_Latest.xlsx"
INDEX_DIR = ROOT / "00_Master" / "index"
STATS_DIR = ROOT / "06_Statistics"

EXPORT_FIELDS = [
    "文献ID", "DOI", "DOI_normalized", "英文题名", "英文题名_normalized",
    "中文题名", "中文题名_normalized", "第一作者", "年份", "期刊/来源",
    "官方网页", "核验状态",
]

PUNCT_TRANSLATION = str.maketrans(
    "", "", string.punctuation + "，。、《》？?！!：:；;（）()【】[]""''\"·—–-_"
)


def normalize_doi(v: Any) -> str:
    t = str(v).strip().lower() if v else ""
    t = re.sub(r"^https?://(dx\.)?doi\.org/", "", t)
    t = re.sub(r"^doi\s*:\s*", "", t, flags=re.IGNORECASE)
    return t.strip()


def normalize_title(v: Any) -> str:
    t = str(v).strip().lower() if v else ""
    t = t.translate(PUNCT_TRANSLATION)
    return re.sub(r"\s+", " ", t).strip()


def extract_records(workbook_path: Path | None = None) -> list[dict[str, str]]:
    """Extract all literature records from the master workbook."""
    src = workbook_path or MASTER
    wb = load_workbook(src, read_only=True, data_only=False)
    try:
        ws = wb["完整文献矩阵"]
        # Find header row (should be row 1)
        headers: dict[str, int] = {}
        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column

        records: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            rid = _cell_val(row, headers, "文献ID")
            if not rid:
                continue

            doi_val = _cell_val(row, headers, "DOI")
            en_val = _cell_val(row, headers, "英文题名")
            zh_val = _cell_val(row, headers, "中文题名")

            records.append({
                "文献ID": rid,
                "DOI": doi_val,
                "DOI_normalized": normalize_doi(doi_val),
                "英文题名": en_val,
                "英文题名_normalized": normalize_title(en_val),
                "中文题名": zh_val,
                "中文题名_normalized": normalize_title(zh_val),
                "第一作者": _cell_val(row, headers, "第一作者"),
                "年份": _cell_val(row, headers, "年份"),
                "期刊/来源": _cell_val(row, headers, "期刊/来源"),
                "官方网页": _cell_val(row, headers, "官方网页"),
                "核验状态": _cell_val(row, headers, "核验状态"),
            })
        return records
    finally:
        wb.close()


def _cell_val(row, headers: dict[str, int], field: str) -> str:
    col = headers.get(field, -1)
    if col < 0:
        return ""
    for cell in row:
        try:
            if cell.column == col:
                return str(cell.value).strip() if cell.value is not None else ""
        except AttributeError:
            continue
    return ""


def export_index(workbook_path: Path | None = None) -> dict[str, Any]:
    """Regenerate master index CSV, JSONL and report. Returns summary dict."""
    records = extract_records(workbook_path)
    src = workbook_path or MASTER
    INDEX_DIR.mkdir(parents=True, exist_ok=True)
    STATS_DIR.mkdir(parents=True, exist_ok=True)

    csv_path = INDEX_DIR / "Deep_CBM_master_index.csv"
    jsonl_path = INDEX_DIR / "Deep_CBM_master_index.jsonl"
    today = date.today().strftime("%Y%m%d")
    report_path = STATS_DIR / f"Master_Index_Report_{today}.md"

    # Write CSV
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=EXPORT_FIELDS)
        w.writeheader()
        w.writerows(records)

    # Write JSONL
    with jsonl_path.open("w", encoding="utf-8", newline="\n") as fh:
        for r in records:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")

    # Compute stats
    doi_missing = sum(1 for r in records if not r["DOI_normalized"])
    doi_dup = _dup_count([r["DOI_normalized"] for r in records])
    en_dup = _dup_count([r["英文题名_normalized"] for r in records])
    zh_dup = _dup_count([r["中文题名_normalized"] for r in records])

    ids: list[int] = []
    for r in records:
        m = re.search(r"DCBM[-_\s]?(\d+)", r["文献ID"], re.IGNORECASE)
        if m:
            ids.append(int(m.group(1)))
    ids_continuous = (sorted(ids) == list(range(min(ids), max(ids) + 1))) if ids else False
    last_id = f"DCBM-{max(ids):03d}" if ids else "N/A"

    report = f"""# Master Index Report

- 运行日期：{date.today().isoformat()}
- 源文件路径：`{src.relative_to(ROOT).as_posix()}`
- 第一张 sheet 名称：完整文献矩阵
- 表头行：1
- 总记录数：{len(records)}
- 最后一个 DCBM 编号：{last_id}
- 文献ID是否连续：{'是' if ids_continuous else '否'}
- DOI 缺失数量：{doi_missing}
- DOI 重复数量：{doi_dup}
- 英文题名重复数量：{en_dup}
- 中文题名重复数量：{zh_dup}
- 缺失字段列表：无

## 输出文件路径

- `{csv_path.relative_to(ROOT).as_posix()}`
- `{jsonl_path.relative_to(ROOT).as_posix()}`
- `{report_path.relative_to(ROOT).as_posix()}`
"""
    with report_path.open("w", encoding="utf-8", newline="\n") as f:
        f.write(report)

    return {
        "record_count": len(records),
        "last_id": last_id,
        "id_continuous": ids_continuous,
        "doi_missing": doi_missing,
        "doi_duplicates": doi_dup,
        "english_title_duplicates": en_dup,
        "chinese_title_duplicates": zh_dup,
        "csv_path": str(csv_path.relative_to(ROOT)),
        "jsonl_path": str(jsonl_path.relative_to(ROOT)),
        "report_path": str(report_path.relative_to(ROOT)),
    }


def _dup_count(items: list[str]) -> int:
    counts = Counter(i for i in items if i)
    return sum(max(0, c - 1) for c in counts.values())
